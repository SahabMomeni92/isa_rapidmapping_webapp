# -*- coding: utf-8 -*-
#!/usr/bin/env python2
########
#ISA Automation Flood Rapid Mapping Version 2.0.0
# Author: sahab momeni (sahabee92@gmail.com)
# version 2.0.0
# using arcpy

version = '2.0.0'
#####

import arcpy

import openpyxl


import os

from arcpy.sa import *


import sys

reload(sys)


import datetime

sys.setdefaultencoding('utf8')

###  mapframe informations


mapframe  = arcpy.mapping.MapDocument(r"C:\flood\MapFrame.mpk")

mdf = arcpy.mapping.ListDataFrames(mapframe,"*")

MapLayers = arcpy.mapping.ListLayers(mapframe,"*",mdf[2])


#import parameter

center_point = r"C:\flood\rfile\center_point.shp"

iran_c = r"C:\flood\shapes\iran_c.shp"


drawMap_id = int(sys.argv[1])



import json

import requests

url = 'http://127.0.0.1:8000/api/expert_panel/drawmap/{}/'.format(drawMap_id)

headers = {
            "content-type":"application/json"
        }
   

r = requests.get(url,headers=headers)    

resp = json.loads(r.text)

title = resp['title']

Flood = resp['tiff_address']

Frame = '#'

Extra_distance = 0

# if Extra_distance == '#' or not Extra_distance:

#     Extra_distance = 0
# else:
#     Extra_distance = float(Extra_distance)

place = MapLayers[0]

landuse = MapLayers[3]

road = MapLayers[1]

Crop_Land = '#'

background_image = '#'

background_imageName = '#'

pahne_area = rad_area = 0

flood_polygone = r"C:\flood\flood\export.gdb\flood"

cropflood_polygone = r"C:\flood\flood\export.gdb\crtpflood"

cropframe_polygone = r"C:\flood\flood\export.gdb\crtpframe"

Export_Output_Intersect = r"C:\flood\flood\export.gdb\eoi"

ebm_flood = r"C:\flood\flood\export.gdb\ebm_flood"

ebm_frame = r"C:\flood\flood\export.gdb\ebm_frame"

exportfolder = r"C:\flood\export"


####################################################################

#functions

def get_ScalebarElementMinWidth(scale_number,W0,Wmin,Wmax):

    L0 = (scale_number * W0 )/ float( 10 ** 5)
    
    N = int(L0/float(4))
    
    if N == 0:
        
        candid_min = 1
        
        candid_max = 4
        
        candid_mean = 2
        
    else:

        candid_min = 4 * N
        
        candid_max = 4 * (N + 1 ) 
        
        candid_mean = (candid_max + candid_min) / float(2) 
        
    Wfit_min = candid_min/float(scale_number)*(10**5)
    
    Wfit_max = candid_max/float(scale_number)*(10**5)
    
    Wfit_mean = candid_mean/float(scale_number)*(10**5)
    
    Smean = 1
    
    Smax = Smin = 0
    
    if Wfit_min >= Wmin and Wfit_min <= Wmax:
    
        Smin = 1
        
        Smean = 0 
        
    if Wfit_max >= Wmin and Wfit_max <= Wmax:

        Smax= 1
        
        Smean = 0 
    
    rating_min = Smin*(abs(L0-Wfit_min))
    
    rating_max = Smax*(abs(L0-Wfit_max))
    
    rating_mean=Smean*(abs(L0-Wfit_max))
    
    if rating_min==0 and rating_max ==0:
          
            mw = (Wfit_mean/float(8))
            
    else:

            if rating_min < rating_max:
            
                mw = (Wfit_min/float(8))
                
            else:

                mw = (Wfit_max/float(8))        

    return mw      

def get_fitScaleNumber(dataFrame):

    sn0 = dataFrame.scale
    
    return (round(sn0/float(1000),0)*1000)

def createTextList(TEXT):

    textList = []

    word = ''

    for part in TEXT:

            if part != u' ':

                word = word + part

            else:

                if word == u'\u0645\u06cc':

                    word = word + u' '

                else:

                    textList.append(word)

                    word = ''

    textList.append(word)

    return textList

def reverse(string):
    
    res = ''
    
    for i in xrange(len(string),0,-1):
        
        res += string[i-1]
        
    return res

def convert_wgs_to_utm(lon, lat):
    
    utm_band = str((math.floor((lon + 180) / 6 ) % 60) + 1)
    
    if len(utm_band) == 1:
        
        utm_band = '0'+utm_band
        
    if lat >= 0:
        
        epsg_code = '326' + utm_band
        
    else:
        
        epsg_code = '327' + utm_band
        
    return epsg_code

def GetUTMZONE(layer,Extra_distance,Order_Frame):
        
        min_x = min_y = 99999999999999999999999999999999999999999999

        max_x = max_y = -99999999999999999999999999999999999999999999

        lat = lon = n = 0

        with arcpy.da.UpdateCursor(layer,"SHAPE@XY") as cursor:
            
            for row in cursor:
            
                lon = row[0][0] + lon
            
                lat = row[0][1] + lat

                n = n + 1      
                        
        lat = lat/float(n)

        lon = lon/float(n)

        zone = int(float(convert_wgs_to_utm(lon, lat)))

        del cursor

        #if Order_Frame == True

        if Order_Frame == True or Order_Frame == False:

                with arcpy.da.UpdateCursor(layer,["SHAPE@","OBJECTID"],spatial_reference=arcpy.SpatialReference(zone)) as cursor:

                        for row in cursor:

                            for part in row[0]:                                

                                    for point in part:

                                          if point:  

                                            if point.X > max_x:

                                                    max_x = point.X

                                            if point.X < min_x:

                                                    min_x = point.X

                                            if point.Y > max_y:

                                                    max_y = point.Y

                                            if point.Y < min_y:

                                                    min_y = point.Y 
        
        return (zone,min_x,max_x,min_y,max_y,lat,lon)

def Create_Frame(GU,Extra_distance):

        array = arcpy.Array([arcpy.Point(GU[1]-Extra_distance, GU[3]-Extra_distance),
                     arcpy.Point(GU[2]+Extra_distance, GU[3]-Extra_distance),
                     arcpy.Point(GU[2]+Extra_distance, GU[4]+Extra_distance),
                     arcpy.Point(GU[1]-Extra_distance, GU[4]+Extra_distance)
                     ])
        
        polygon = arcpy.Polygon(array)

        arcpy.CreateFeatureclass_management(r"C:\flood\rfile","frame","POLYGON",spatial_reference=arcpy.SpatialReference(GU[0])) 

        cursor = arcpy.da.InsertCursor(r"C:\flood\rfile\Frame.shp", ['SHAPE@'])

        cursor.insertRow([polygon])

        del cursor

        return r"C:\flood\rfile\Frame.shp"

def file_path(exportfile,name):

    return os.path.join(exportfile,name)

def getRoad_info(layer,road):

    way1 = way2 = way3 = way4 = way5 = 0

    arcpy.Intersect_analysis([layer,road], Export_Output_Intersect, "ALL", "", "INPUT")

    with arcpy.da.SearchCursor(Export_Output_Intersect,['SHAPE@LENGTH','fclass'],spatial_reference=arcpy.SpatialReference(zone)) as cursor:

        for row in cursor:

            if row[1] == 'trunk' or row[1] == 'trunk_link' or row[1] == 'primary' or row[1] == 'primary_link' or row[1] == 'motorway' or row[1] == 'motorway_link':

                way1 = way1 + row[0]

            elif row[1] == 'secondary' or row[1] == 'secondary_link':

                way2 = way2 + row[0]

            elif row[1] == 'track_grade5' or row[1] == 'track_grade4' or row[1] == 'track_grade3' or row[1] == 'track_grade2' or row[1] == 'track_grade1' or row[1] ==  'track' or row[1] ==  'unclassified':

                way3 = way3 + row[0]

            elif row[1] == 'residential' or row[1] == 'service':

                way4 = way4 + row[0]

            elif row[1] == 'tertiary' or row[1] == 'tertiary_link':

                way5 = way5 + row[0]    

    arcpy.Delete_management(Export_Output_Intersect)             

    return (way1,way2,way3,way4,way5)

def Landuse_info(layer,landuse,crop,Crop_Order = True):

    farm = other = 0

    arcpy.Intersect_analysis([layer,landuse], Export_Output_Intersect, "ALL", "", "INPUT")

    with arcpy.da.SearchCursor(Export_Output_Intersect,['SHAPE@AREA','fclass'],spatial_reference=arcpy.SpatialReference(zone)) as cursor:

        for row in cursor:

            if row[1] != 'farm':

                    other = other + row[0]

            else:

                    farm = farm + row[0]
                    
    arcpy.Delete_management(Export_Output_Intersect)

    if Crop_Order == True:

        farm = 0

        arcpy.Intersect_analysis([layer,crop], Export_Output_Intersect, "ALL", "", "INPUT")

        with arcpy.da.SearchCursor(Export_Output_Intersect,['SHAPE@AREA','gridcode'],spatial_reference=arcpy.SpatialReference(zone)) as cursor:

            for row in cursor:

                if row[1] == 1:

                    farm = farm + row[0]

        arcpy.Delete_management(Export_Output_Intersect)                

    return (farm,other)

def Place_info(layer,place):

    town = hamlet = village = pop = 0

    arcpy.Intersect_analysis([layer,place], Export_Output_Intersect, "ALL", "", "INPUT")

    with arcpy.da.SearchCursor(Export_Output_Intersect,['population','fclass'],spatial_reference=arcpy.SpatialReference(zone)) as cursor:

        for row in cursor:

            pop = pop + row[0]

            if row[1] == "town" or row[1] == "suburb" or row[1] == "national_capital":

                town = town + 1

            elif row[1] == "village" :

                village = village + 1

            elif row[1] == "hamlet":

                hamlet = hamlet + 1
 
    arcpy.Delete_management(Export_Output_Intersect)

    return (pop,town,village,hamlet)

def gregorian_to_jalali(gy,gm,gd):
 g_d_m=[0,31,59,90,120,151,181,212,243,273,304,334]
 
 if(gy>1600):
         
  jy=979
  
  gy-=1600
  
 else:
         
  jy=0
  
  gy-=621
  
 if(gm>2):
         
  gy2=gy+1
  
 else:
         
  gy2=gy
  
 days=(365*gy) +(int((gy2+3)/4)) -(int((gy2+99)/100)) +(int((gy2+399)/400)) -80 +gd +g_d_m[gm-1]
 
 jy+=33*(int(days/12053))
 
 days%=12053
 
 jy+=4*(int(days/1461))
 
 days%=1461
 
 if(days > 365):
         
  jy+=int((days-1)/365)
  
  days=(days-1)%365
  
 if(days < 186):
         
  jm=1+int(days/31)
  
  jd=1+(days%31)
  
 else:
         
  jm=7+int((days-186)/30)
  
  jd=1+((days-186)%30)
  
 return [jy,jm,jd]
    
                                                
def Check_gdbSpace(path):

    if arcpy.Exists(path):

            arcpy.Delete_management(path)
        

##########################################################################################################################################################

# delete last flood files

Check_gdbSpace(ebm_flood)

Check_gdbSpace(ebm_frame)

Check_gdbSpace(flood_polygone)

Check_gdbSpace(cropflood_polygone)

Check_gdbSpace(cropframe_polygone)

Check_gdbSpace(r"C:\flood\rfile\Frame.shp")

Check_gdbSpace(r"C:\flood\rfile\center_point.shp")

Check_gdbSpace(r"C:\flood\rfile\ostan_selection.shp")

Check_gdbSpace(r"C:\flood\rfile\center_roosta.shp")

Check_gdbSpace(r"C:\flood\rfile\roosta.shp")

Check_gdbSpace(r"C:\flood\rfile\iranc_selection.shp")

Check_gdbSpace(Export_Output_Intersect)

mxdpath = file_path(exportfolder,"floodmap.mxd")

Check_gdbSpace(mxdpath)

if arcpy.Exists(r"C:\flood\flood\export.gdb") == False:

        arcpy.CreateFileGDB_management(r"C:\flood\flood","export")

        print "yes"

        #arcpy.env.workspace = r"C:\flood\flood\export.gdb"

#else:
        #arcpy.env.workspace = r"C:\flood\flood\export.gdb"

###############################################################################################

#Create Flood and stats
        
arcpy.RasterToPolygon_conversion(Flood,flood_polygone, "SIMPLIFY", "Value")

with arcpy.da.UpdateCursor(flood_polygone,"gridcode") as cursor:
    
     for row in cursor:
         
         if row[0] == 0:
             
            cursor.deleteRow() 

if Frame == '#' or not Frame:

        GU =  GetUTMZONE(flood_polygone,Extra_distance,Order_Frame = True)

        zone = GU[0]

        #lat = GU[5]

        #lon = GU[6]

        Frame = Create_Frame(GU,Extra_distance)

else:

        GU =  GetUTMZONE(Frame,Extra_distance,Order_Frame = False)
        
        zone = GU[0]

        #lat = GU[5]

        #lon = GU[6]

#zone = 32639

array = (((GU[1]+GU[2])*0.5),((GU[3]+GU[4])*0.5))

point = arcpy.Point(array[0],array[1]) 

arcpy.CreateFeatureclass_management(r"C:\flood\rfile","center_point","POINT",spatial_reference=arcpy.SpatialReference(zone))

pcursor = arcpy.da.InsertCursor(r"C:\flood\rfile\center_point.shp", ['SHAPE@'])

pcursor.insertRow([point])

del pcursor

if arcpy.Exists(Crop_Land):

    ebmframe = ExtractByMask(Crop_Land,Frame)

    ebmframe.save(ebm_frame)

    ebmflood = ExtractByMask(Crop_Land,flood_polygone)

    ebmflood.save(ebm_flood)

    arcpy.RasterToPolygon_conversion(ebmframe,cropframe_polygone, "SIMPLIFY", "Value")

    arcpy.RasterToPolygon_conversion(ebmflood,cropflood_polygone, "SIMPLIFY", "Value")

    with arcpy.da.UpdateCursor(cropflood_polygone,'gridcode') as cursor:

            for row in cursor:

                    if row[0] == 0:

                            cursor.deleteRow()

    with arcpy.da.UpdateCursor(cropframe_polygone,'gridcode') as cursor:

            for row in cursor:

                    if row[0] == 0:

                            cursor.deleteRow()                        

    Crop_Order = True

else:

    cropframe_polygone = cropflood_polygone = 0

    Crop_Order = False
    
with arcpy.da.SearchCursor(flood_polygone,['SHAPE@AREA','gridcode'],spatial_reference=arcpy.SpatialReference(zone)) as cursor:
    
     for row in cursor:

         if row[1] == 1:

             pahne_area = pahne_area + row[0]

         elif row[1] ==2:

             rad_area = rad_area + row[0]
              
             
flood_intersect_road = getRoad_info(flood_polygone,road)

frame_intersect_road = getRoad_info(Frame,road)

flood_intersect_landuse = Landuse_info(flood_polygone,landuse,cropflood_polygone,Crop_Order)

frame_intersect_landuse = Landuse_info(Frame,landuse,cropframe_polygone,Crop_Order)

flood_intersect_place = Place_info(flood_polygone,place)

frame_intersect_place = Place_info(Frame,place)
         
#arcpy.Delete_management(flood_polygone)

arcpy.Delete_management(cropflood_polygone)

arcpy.Delete_management(cropframe_polygone) 

arcpy.Delete_management(ebm_frame)

arcpy.Delete_management(ebm_flood)

#######################################################################################################

### create map element list

dates = {}

stats = {}

# set data informations elements in list

# workbook = openpyxl.load_workbook(entered_informations)

# sheet = workbook.active

today = datetime.date.today()
                
pt = gregorian_to_jalali(today.year,today.month,today.day) # persian today

dates[1] = resp['disaster_date']

dates[2] = resp['image_date']

dates[3] = resp['proccess_date']

dates[4] = str(pt[0]) + "/" + str(pt[1]) + "/" + str(pt[2]) 

sat_name = resp['sat']

#### set stats table informations elements in list

stats[1] = round(pahne_area /float(10**6),2)

stats[2] = round(rad_area /float(10**6),2)

stats[3] = resp['total_pop']

stats[4] = round(frame_intersect_place[0],0)

stats[5] = resp['total_city']

stats[6] = round(frame_intersect_place[1],0)

stats[7] = resp['total_vilage']

stats[8] = round(frame_intersect_place[2],0)

stats[9] = resp['total_hamlet']

stats[10] = round(frame_intersect_place[3],0)

stats[11] = round(flood_intersect_road[0]/float(1000),2)

stats[12] = round(frame_intersect_road[0]/float(1000),2)

stats[13] = round(flood_intersect_road[1]/float(1000),2)

stats[14] = round(frame_intersect_road[1]/float(1000),2)

stats[15] = round(flood_intersect_road[4]/float(1000),2)

stats[16] =  round(frame_intersect_road[4]/float(1000),2)

stats[17] =  round(flood_intersect_road[2]/float(1000),2)

stats[18] =  round(frame_intersect_road[2]/float(1000),2)

stats[19] =  round(flood_intersect_road[3]/float(1000),2)

stats[20] =  round(frame_intersect_road[3]/float(1000),2)

stats[21] =  round(flood_intersect_landuse[0] / float(10**4),2)

stats[22] =  round(frame_intersect_landuse[0] / float(10**4),2)

stats[23] =  round(flood_intersect_landuse[1]/ float(10**4),2)

stats[24] =  round(frame_intersect_landuse[1]/ float(10**4),2)

##### set elements on mapFrame

for elm in arcpy.mapping.ListLayoutElements(mapframe, "TEXT_ELEMENT"):

    if elm.name == 'zone':
                
                elm.text = str(zone - 32600) + "N"
                
    elif elm.name.startswith('date0'):
    
                index = int(elm.name[-1])

                elm.text = dates[index] 
                
    elif elm.name.startswith('img0'):

                if elm.name[-1] == '1':
                
                        elm.text = sat_name
                        
                if elm.name[-1] == '2' and background_imageName == True:
                
                        elm.text = background_imageName 
                                                          
    elif elm.name.startswith('selm0'):
    
                index = int(elm.name[-1])
                
                elm.text = stats[index] 
                
    elif elm.name.startswith('selm') and not elm.name.startswith('selm0'):

                index = int(elm.name[-2:])
                
                elm.text = stats[index]
                
    elif elm.name == 'title':
    
                title01 = u'\u0646\u0642\u0634\u0647 \u067e\u0647\u0646\u0647 \u0633\u06cc\u0644\u0627\u0628'
                
                title02 = u'\u062f\u0631 \u062a\u0627\u0631\u06cc\u062e'

                elm.text = title01 + " " + title + " " + title02 + " " + dates[4]
                
######################################################################################

# add flood and frame to Map and minimap Layers

newlayer = arcpy.mapping.Layer(Frame)

arcpy.mapping.AddLayer(mdf[2], newlayer, "TOP")

arcpy.mapping.AddLayer(mdf[0], newlayer, "TOP")

newlayer = arcpy.mapping.Layer(Flood)

arcpy.mapping.AddLayer(mdf[2], newlayer, "TOP")

# set layers

overviewmapLayer =  arcpy.mapping.ListLayers(mapframe,"*",mdf[0])

MapLayers = arcpy.mapping.ListLayers(mapframe,"*",mdf[2])

med_overviewmapLayer =  arcpy.mapping.ListLayers(mapframe,"*",mdf[3])

# calculate county and province of POI

arcpy.Intersect_analysis([center_point, overviewmapLayer[3]], Export_Output_Intersect, "ALL", "", "INPUT")

#newlayer = arcpy.mapping.Layer(Export_Output_Intersect)

#arcpy.mapping.AddLayer(mdf[0], newlayer, "BOTTOM")

overviewmapLayer = arcpy.mapping.ListLayers(mapframe, "*", mdf[0])

#mapframe.saveACopy(file_path(exportfolder,"floodmap.mxd"))

with arcpy.da.SearchCursor(Export_Output_Intersect, ['pname']) as cursor:
    
    fq = 1

    for row in cursor:

        ostan = row[0]

        if fq == 1:

            qurey = """ Name = '%s'""" % ostan.decode('utf-8')

            fq = 0
        else:

            qurey = str(qurey) + " AND " + """ Name = '%s'""" % ostan.decode('utf-8')

    arcpy.SelectLayerByAttribute_management(overviewmapLayer[4], "NEW_SELECTION", qurey)
        

arcpy.CreateFeatureclass_management(r"C:\flood\rfile", "ostan_selection", "POLYGON",
                                    spatial_reference=arcpy.SpatialReference(zone))

arcpy.AddField_management(r"C:\flood\rfile\ostan_selection.shp", "Name", "TEXT", field_is_nullable="NULLABLE")

with arcpy.da.SearchCursor(overviewmapLayer[4], ['SHAPE@', 'Name']) as cursor:

    for row in cursor:

        with arcpy.da.InsertCursor(r"C:\flood\rfile\ostan_selection.shp", ["SHAPE@", "Name"]) as Icursor:

            Icursor.insertRow([row[0], row[1]])

        del Icursor

arcpy.SelectLayerByAttribute_management(overviewmapLayer[4], "CLEAR_SELECTION")

overviewmapLayer = arcpy.mapping.ListLayers(mapframe, "*", mdf[0])

with arcpy.da.SearchCursor(Export_Output_Intersect, ['pname', 'Name', 'ccenter_na']) as cursor:

    for row in cursor:

        ostan = row[0]

        roosta = row[1]

        croosta = row[2]

qurey = """ pname = '%s'""" % ostan.decode('utf-8')

arcpy.SelectLayerByAttribute_management(overviewmapLayer[3], "NEW_SELECTION", qurey)

arcpy.CreateFeatureclass_management(r"C:\flood\rfile", "iranc_selection", "POLYGON",
                                    spatial_reference=arcpy.SpatialReference(zone))

arcpy.AddField_management(r"C:\flood\rfile\iranc_selection.shp", "Name", "TEXT", field_is_nullable="NULLABLE")

with arcpy.da.SearchCursor(overviewmapLayer[3], ['SHAPE@', 'Name']) as cursor:

    for row in cursor:

        with arcpy.da.InsertCursor(r"C:\flood\rfile\iranc_selection.shp", ["SHAPE@", "Name"]) as Icursor:

            Icursor.insertRow([row[0], row[1]])

        del Icursor

arcpy.SelectLayerByAttribute_management(overviewmapLayer[3], "CLEAR_SELECTION")

arcpy.Delete_management(Export_Output_Intersect)

# add layer to med_minimap

newlayer = arcpy.mapping.Layer(r"C:\flood\rfile\ostan_selection.shp")

arcpy.mapping.AddLayer(mdf[3],newlayer,"TOP")

newlayer = arcpy.mapping.Layer(r"C:\flood\rfile\iranc_selection.shp")

arcpy.mapping.AddLayer(mdf[3],newlayer,"TOP")

newlayer = arcpy.mapping.Layer(center_point)

arcpy.mapping.AddLayer(mdf[3],newlayer,"TOP")

med_overviewmapLayer =  arcpy.mapping.ListLayers(mapframe,"*",mdf[3])

# Map Layer zoom

Extent = MapLayers[1].getExtent(True) # zood to frame

mdf[2].extent = Extent

mdf[2].scale = mdf[2].scale * 1.25

fitScale = get_fitScaleNumber(mdf[2])

mdf[2].scale = fitScale

for elm in arcpy.mapping.ListLayoutElements(mapframe, "TEXT_ELEMENT"):

    if elm.name == 'scale':
    
        scale_text = '1:' + str(fitScale) 
    
        elm.text = scale_text[:-2] 

# mini Map Layer zoom

overviewmaplayer = arcpy.mapping.ListLayers(mapframe,"*",mdf[0])

Extent = MapLayers[0].getExtent(True) # zood to frame

mdf[0].extent = Extent

mdf[0].scale = mdf[0].scale * 1.25

# Med Map Layer zoom

Extent = med_overviewmapLayer[1].getExtent(True) # zood to frame

mdf[3].extent = Extent

mdf[3].scale = mdf[3].scale * 1.25

# ApplySymbology to new layers



if rad_area>0:

    arcpy.ApplySymbologyFromLayer_management(MapLayers[0],r"C:\flood\layers\floodwithrad.lyr")

    newlayer = arcpy.mapping.Layer(flood_polygone)

    arcpy.mapping.AddLayer(mdf[2],newlayer,"TOP")

    MapLayers = arcpy.mapping.ListLayers(mapframe,"*",mdf[2])

    arcpy.ApplySymbologyFromLayer_management(MapLayers[0],r"C:\flood\layers\rad.lyr")

    arcpy.ApplySymbologyFromLayer_management(MapLayers[2],r"C:\flood\layers\Frame.lyr")
    
else:

    arcpy.ApplySymbologyFromLayer_management(MapLayers[0],r"C:\flood\layers\flood.lyr")

    arcpy.ApplySymbologyFromLayer_management(MapLayers[1],r"C:\flood\layers\Frame.lyr")
 
med_overviewmapLayer[1].showLabels = True
 
arcpy.ApplySymbologyFromLayer_management(overviewmapLayer[0],r"C:\flood\layers\overviewmapframe.lyr")

arcpy.ApplySymbologyFromLayer_management(med_overviewmapLayer[2],r"C:\flood\layers\iran_ostan selection.lyr")

arcpy.ApplySymbologyFromLayer_management(med_overviewmapLayer[1],r"C:\flood\layers\iran_c selection.lyr")

arcpy.ApplySymbologyFromLayer_management(med_overviewmapLayer[0],r"C:\flood\layers\center_point.lyr")

arcpy.mapping.UpdateLayer(mdf[3],med_overviewmapLayer[1]
,arcpy.mapping.Layer(r"C:\flood\layers\iran_c selection.lyr"), True)


# Set graphic elemenst

for elm in arcpy.mapping.ListLayoutElements(mapframe, "PICTURE_ELEMENT"):
        
    if elm.name == "stats":
            
        elm.sourceImage = file_path(exportfolder,"Flood Damage_report.png")

        elm.elementHeight = 5.3621

        elm.elementWidth = 9.5

        elm.elementPositionX = 36.5

        elm.elementPositionY = 7.918

    if elm.name == "isa logo":

        elm.sourceImage = r"C:\flood\isa logo.jpg"

    if elm.name == "legend":

        elm.sourceImage = r"C:\flood\plz3.jpg"

#### set discriptions Texts

dworkbook = openpyxl.load_workbook(r"C:\flood\map_discription.xlsx")

dsheet = dworkbook.active

t1 = u'\u0627\u06cc\u0646 \u0646\u0642\u0634\u0647 \u067e\u0647\u0646\u0647 \u0633\u06cc\u0644\u0627\u0628\u06cc'

#t1 = 'این نقشه پهنه سیلابی'

#t2 = 'را نشان می دهد. بر اساس تحلیل صورت گرفته در محدوده مورد مطالعه ( کادر سبز)  مساحت کل پهنه سیلاب برابر با'

t2 = u'\u0631\u0627 \u0646\u0634\u0627\u0646 \u0645\u06cc \u062f\u0647\u062f. \u0628\u0631 \u0627\u0633\u0627\u0633 \u062a\u062d\u0644\u06cc\u0644 \u0635\u0648\u0631\u062a \u06af\u0631\u0641\u062a\u0647 \u062f\u0631 \u0645\u062d\u062f\u0648\u062f\u0647 \u0645\u0648\u0631\u062f \u0645\u0637\u0627\u0644\u0639\u0647 ( \u06a9\u0627\u062f\u0631 \u0633\u0628\u0632)  \u0645\u0633\u0627\u062d\u062a \u06a9\u0644 \u067e\u0647\u0646\u0647 \u0633\u06cc\u0644\u0627\u0628 \u0628\u0631\u0627\u0628\u0631 \u0628\u0627'

#t3 = 'کیلومتر مربع اندازه گیری شده است و تخمین زده می شود حدود'.decode("UTF8")

t3 = u'\u06a9\u06cc\u0644\u0648\u0645\u062a\u0631 \u0645\u0631\u0628\u0639 \u0627\u0646\u062f\u0627\u0632\u0647 \u06af\u06cc\u0631\u06cc \u0634\u062f\u0647 \u0627\u0633\u062a \u0648 \u062a\u062e\u0645\u06cc\u0646 \u0632\u062f\u0647 \u0645\u06cc \u0634\u0648\u062f \u062d\u062f\u0648\u062f'

#t4 = 'کیلومتر از راه های مواصلاتی'

t4 = u'\u06a9\u06cc\u0644\u0648\u0645\u062a\u0631 \u0627\u0632 \u0631\u0627\u0647 \u0647\u0627\u06cc \u0645\u0648\u0627\u0635\u0644\u0627\u062a\u06cc'

#t5 = 'هکتار از زمین های کشاورزی'

t5 = u'\u0647\u06a9\u062a\u0627\u0631 \u0627\u0632 \u0632\u0645\u06cc\u0646 \u0647\u0627\u06cc \u06a9\u0634\u0627\u0648\u0631\u0632\u06cc'

#t6 = 'منطقه  شهری'

t6 = u'\u0645\u0646\u0637\u0642\u0647  \u0634\u0647\u0631\u06cc'

#t7 = 'منطقه روستایی و بیش از '

t7 = u'\u0645\u0646\u0637\u0642\u0647 \u0631\u0648\u0633\u062a\u0627\u06cc\u06cc \u0648 \u0628\u06cc\u0634 \u0627\u0632 '

#t8 = 'هزار نفر جمعیت درگیر این سیلاب  هستند'.decode("UTF8")

t8 = u'\u0647\u0632\u0627\u0631 \u0646\u0641\u0631 \u062c\u0645\u0639\u06cc\u062a \u062f\u0631\u06af\u06cc\u0631 \u0627\u06cc\u0646 \u0633\u06cc\u0644\u0627\u0628  \u0647\u0633\u062a\u0646\u062f'

defaultText = t1 + " " + str(title) + " " + t2 + " " +  str(stats[1]) + " " + t3 + " " + str(stats[13] + stats[15] + stats[17] + stats[19] +stats[11]) + " " + t4 + " " + str(stats[21]) + " " + t5 + " " + str(stats[5]) + " " + t6 + " " + str(stats[7]) + " " + t7 + " " + str(round(stats[3]/float(1000),0)) + " " + t8 + '.'
 
dsheet.cell(row=1, column=1).value  = defaultText
 
dworkbook.save(r"C:\flood\map_discription.xlsx")   

should_restart = True 

while should_restart == True:

    should_restart = False

    os.system(r"C:\flood\map_discription.xlsx")

    dworkbook = openpyxl.load_workbook(r"C:\flood\map_discription.xlsx")

    dsheet = dworkbook.active

    TEXT = dsheet.cell(row=1, column=1).value 

    for elm in arcpy.mapping.ListLayoutElements(mapframe, "TEXT_ELEMENT"):

        if elm.name == "text1":

            mainText = elm

        if elm.name == "text2":

            lastLine = elm

    newline = ''
    
    pre_line = ""

    finalText = ''

    fitText = ''

    LineCount = 1

    textList = createTextList(TEXT)

    for word in textList:

        if newline == '':

            newline = word
            
            pre_word = word          
            
        else:

            testText = newline + ' ' + word

            mainText.text = testText

            mainText.fontSize = 16

            if mainText.elementWidth <= 19.8:

                    newline = testText
                    
                    pre_word = word
                    
                    line_length = len(newline)
                    
                    word_length = len(word) + 1 
                                                   
            else:
                    LineCount = LineCount + 1                           
                    
                    if LineCount == 6:

                        should_restart = True
                        
                        dsheet.cell(row=1, column=1).value  = defaultText
                        
                        break
                        
                    if fitText == "":

                           fitText = newline 
                           
                           newline = word
                           
                           pre_word = word
                    else:
                   
                           testText = fitText + '\n' + newline 
                    
                           mainText.text = testText
                    
                           if mainText.elementWidth <= 19.8:
                           
                               fitText = testText
                    
                               pre_word = word

                               newline = word
                               
                           else:
                                
                               newline = newline[:(line_length-word_length)]
                               
                               fitText = fitText + '\n' + newline 
                               
                               newline = pre_word + " " + word
                               

mainText.text = fitText
                    
                    
if LineCount == 1:

    lastLine.text = ""

    #mainText.text = finalText

    mainText.elementPositionX = lastLine.elementPositionX = 82

else:

    LineCount = LineCount + 1

    xl = ((5-LineCount)*mainText.elementHeight)/float(2)     
    
    mainText.elementPositionY = 29.2749 - xl

    mainText.elementPositionX =  82

    #lastLine.elementPositionX =  mainText.elementPositionX + mainText.elementWidth/float(2) #- lastLine.elementWidth/float(2)
    
    lastLine.elementPositionX = mainText.elementPositionX

    #lastLine.elementPositionY = mainText.elementPositionY - mainText.elementHeight/float(2) - lastLine.elementHeight/float(2)
    
    lastLine.elementPositionY = mainText.elementPositionY - mainText.elementHeight

    lastLine.text = newline
    
    
#### scale bar

#sb_len = get_ScalebarElementMinWidth(scale_number,W0,Wmin,Wmax)

maxWidth = 12.8515

minWidth = 6.676

W0 = 9.5

sb_len = get_ScalebarElementMinWidth(fitScale,W0,minWidth,maxWidth)

#sb_len = 1

sb_height = 0.352778

sb_px = 62.2 + sb_len

sb_py = 40.0768

scaleBar = {}

scaleBar[0] = (7*sb_len,sb_height,62.2 + 7*sb_len,sb_py)

scaleBar[1] = (sb_len,sb_height + 0.0176389,sb_px,sb_py)

scaleBar[2] = (sb_len,sb_height,sb_px + sb_len,sb_py)

scaleBar[3] = (2*sb_len,sb_height + 0.0176389,sb_px +  3*sb_len,sb_py)

scaleBar[4] = (2*sb_len,sb_height,sb_px + 5*sb_len,sb_py)

scaleBar[5] = (2*sb_len,sb_height + 0.0176389,sb_px + 7*sb_len,sb_py)

for elm in arcpy.mapping.ListLayoutElements(mapframe, "TEXT_ELEMENT"):

    if elm.name.startswith('sb_elm'):
    
#        if elm.name.endswith('0') or elm.name.endswith('2') or elm.name.endswith('4'):

            index = int(elm.name[-1])
        
            elm.elementWidth,elm.elementHeight,elm.elementPositionX,elm.elementPositionY = scaleBar[index]
        
    elif elm.name.startswith('sb_unit'):

        elm.elementPositionX,elm.elementPositionY = sb_px + 7*sb_len + 0.105833, sb_py - sb_height
    
    elif elm.name.startswith('sb_num'):
    
        index = int(elm.name[-1])
        
        if index != 0:
        
        
            num_text = (0,sb_len*fitScale*0.00001,2*sb_len*fitScale*0.00001,4*sb_len*fitScale*0.00001,6*sb_len*fitScale*0.00001,8*sb_len*fitScale*0.00001)
    
            elm.elementPositionX,elm.elementPositionY,elm.text = scaleBar[index][2] - 0.5*elm.elementWidth , scaleBar[index][3] + 0.105833 + 0.5*scaleBar[index][1],num_text[index]
    
    #### set version
    
    elif elm.name.startswith('version'):
    
        elm.text = 'ISA Automation Flood Rapid Mapping Version ' + str(version)
            
#### set background_image

if background_image == True:

    addLayer = arcpy.mapping.Layer(satimage)

    arcpy.mapping.AddLayer(mdf[2],addLayer,"BOTTOM")

#basemapLayer = arcpy.mapping.Layer(r"C:\flood\layers\OSMbasemap.lyr")

#arcpy.mapping.AddLayer(overviewmapLayer, basemapLayer, "BOTTOM")
                                    
arcpy.RefreshActiveView()
                                    
mapframe.saveACopy(file_path(exportfolder,"floodmap.mxd"))

mxd = arcpy.mapping.MapDocument(file_path(exportfolder,"floodmap.mxd"))

arcpy.mapping.ExportToJPEG(mxd,file_path(exportfolder,"floodmap.jpg"),resolution = 300)
                                
del mapframe

os.system(file_path(exportfolder,"floodmap.jpg"))






